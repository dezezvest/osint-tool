import os
import csv
import re
import time
from pathlib import Path
from typing import List, Dict, Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None


class FastSearch:
    def __init__(self, data_dir: str = "data"):
        self.data_dir = Path(data_dir)
        self.files = []
        self._scan_files()

    def _scan_files(self):
        if not self.data_dir.exists():
            return
        self.files = list(self.data_dir.glob("*.*"))
        print(f"[+] Найдено файлов: {len(self.files)}")

    def search(self, query: str) -> List[Dict]:
        results = []
        query_clean = query.strip().lower()
        total_files = len(self.files)

        print(f"\n🔍 Поиск: {query}")
        print("-" * 50)

        start_time = time.time()

        for idx, file in enumerate(self.files, 1):
            progress = f"[{idx}/{total_files}]"
            print(f"\r   {progress} Поиск в: {file.name:<40}", end="", flush=True)

            found = self._search_in_file(file, query_clean)
            if found:
                results.extend(found)
                print(f"\r   {progress} ✅ Найдено в: {file.name:<40}")

        elapsed = time.time() - start_time
        print(f"\n   ⏱️ Время: {elapsed:.2f} сек")
        print("-" * 50)

        return results

    def _search_in_file(self, file: Path, query: str) -> List[Dict]:
        ext = file.suffix.lower()
        results = []

        try:
            if ext in ['.csv', '.txt']:
                results = self._search_csv(file, query)
            elif ext == '.xlsx':
                results = self._search_xlsx(file, query)
        except:
            pass

        return results

    def _search_csv(self, file: Path, query: str) -> List[Dict]:
        for delimiter in [',', ';', '\t', '|']:
            try:
                with open(file, 'r', encoding='utf-8-sig', errors='ignore') as f:
                    sample = f.read(4096)
                    f.seek(0)
                    if delimiter in sample:
                        reader = csv.DictReader(f, delimiter=delimiter)
                        if reader.fieldnames:
                            break
            except:
                continue
        else:
            return []

        results = []
        try:
            with open(file, 'r', encoding='utf-8-sig', errors='ignore') as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                for row in reader:
                    for key, value in row.items():
                        if value and query in str(value).lower():
                            results.append({
                                'file': file.name,
                                'data': row
                            })
                            break
        except:
            pass

        return results

    def _search_xlsx(self, file: Path, query: str) -> List[Dict]:
        if openpyxl is None:
            return []

        results = []
        try:
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
            ws = wb.active

            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else f"col_{len(headers)}")

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = str(val).strip() if val is not None else ""

                for key, value in row_dict.items():
                    if value and query in str(value).lower():
                        results.append({
                            'file': file.name,
                            'data': row_dict
                        })
                        break

            wb.close()
        except:
            pass

        return results


def format_result(result: Dict, index: int) -> str:
    lines = []
    lines.append(f"\n📌 [{index}] Источник: {result['file']}")
    lines.append("-" * 40)

    data = result['data']
    for key, value in data.items():
        if value and value.strip():
            lines.append(f"   {key}: {value}")

    lines.append("-" * 40)
    return "\n".join(lines)


def main():
    print("\n" + "=" * 60)
    print("  OSINT TOOL - Поиск по базам")
    print("=" * 60)

    searcher = FastSearch("data")

    if not searcher.files:
        print("\n[!] Нет файлов в папке data!")
        print("[!] Поместите файлы с базами в папку data")
        return

    print("\n" + "=" * 60)
    print("  Вводите что угодно для поиска:")
    print("  - Телефон: 79529637711")
    print("  - Имя: Евгений")
    print("  - Email: test@mail.ru")
    print("  - Username: evgeny")
    print("  - ID: 1485647396")
    print("  - Любое другое слово")
    print("")
    print("  Команды: stats, exit")
    print("=" * 60)

    while True:
        try:
            cmd = input("\n> ").strip()
            if not cmd:
                continue

            if cmd.lower() == "exit":
                break

            if cmd.lower() == "stats":
                print(f"\n📊 Файлов: {len(searcher.files)}")
                total_size = 0
                for f in searcher.files:
                    size = f.stat().st_size / (1024 * 1024)
                    total_size += size
                    print(f"   {f.name} ({size:.1f} MB)")
                print(f"\n   Общий размер: {total_size:.1f} MB")
                continue

            results = searcher.search(cmd)

            if not results:
                print("\n   ❌ Не найдено")
            else:
                print(f"\n✅ Найдено: {len(results)} записей")
                for i, res in enumerate(results[:50], 1):
                    print(format_result(res, i))

                if len(results) > 50:
                    print(f"\n... и ещё {len(results) - 50} записей")

        except KeyboardInterrupt:
            print("\n[!] Выход...")
            break
        except Exception as e:
            print(f"[-] Ошибка: {e}")


if __name__ == "__main__":
    main()
